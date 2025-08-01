import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font

st.set_page_config(page_title="USDA产量对比助手", layout="centered")

st.title("🌾 USDA主要作物产量同比环比一键表格")

# 1. 上传数据区
st.markdown("#### 1. 上传最新两个月的csv（格式同USDA原表，无需修改）")
with st.container():
    col1, col2 = st.columns(2)
    file_july = col1.file_uploader("上传本月csv", type="csv", key="july")
    file_june = col2.file_uploader("上传上月csv", type="csv", key="june")

# 2. 默认大豆/自动扫描
default_commodity_list = ["Oilseed, Soybean"]
default_cn_dict = {"Oilseed, Soybean": "大豆"}

commodity_list = default_commodity_list.copy()
if file_july:
    df_july_tmp = pd.read_csv(file_july)
    commodity_list = sorted(df_july_tmp['Commodity_Description'].dropna().unique())
    file_july.seek(0)  # 重新指针，否则后面再读取会报错

default_index = 0
for i, c in enumerate(commodity_list):
    if c == "Oilseed, Soybean":
        default_index = i
        break

st.markdown("#### 2. 选择作物并填写中文")
cols = st.columns([3, 2])
selected_commodity = cols[0].selectbox(
    "请选择要对比的作物（英文）", commodity_list, index=default_index, label_visibility="collapsed"
)
default_cn = "大豆" if selected_commodity == "Oilseed, Soybean" else ""
cn_commodity = cols[1].text_input("请输入对应的中文名", value=default_cn, label_visibility="collapsed", placeholder="如：大豆")

# 3. 国家名单与编辑
default_countries = [
    {"en": "Brazil", "cn": "巴西"},
    {"en": "Argentina", "cn": "阿根廷"},
    {"en": "Paraguay", "cn": "巴拉圭"},
    {"en": "United States", "cn": "美国"},
    {"en": "China", "cn": "中国"}
]
if "edit_country_data" not in st.session_state:
    df0 = pd.DataFrame(default_countries)
    df0["del"] = False
    st.session_state.edit_country_data = df0

st.markdown("#### 3. 国家名单（可编辑）")
data_local = st.session_state.edit_country_data.copy()
data = st.data_editor(
    data_local,
    use_container_width=True,
    column_order=["en", "cn", "del"],
    hide_index=True,
    key="edit_country"
)

# 添加新国家区域
with st.container():
    st.markdown("#### 添加新国家")
    add_cols = st.columns([4, 4, 2])
    with add_cols[0]:
        new_en = st.text_input("英文名", key="add_en", label_visibility="collapsed", placeholder="英文名")
    with add_cols[1]:
        new_cn = st.text_input("中文名", key="add_cn", label_visibility="collapsed", placeholder="中文名")
    with add_cols[2]:
        st.write("")  # 占位对齐
        if st.button("➕ 添加", key="add_btn", help="添加到国家名单"):
            if new_en.strip() and new_cn.strip():
                new_row = pd.DataFrame([{"en": new_en.strip(), "cn": new_cn.strip(), "del": False}])
                st.session_state.edit_country_data = pd.concat([data, new_row], ignore_index=True)
                st.rerun()
            else:
                st.warning("请填写完整的英文和中文名称！")

    btn_cols = st.columns([1, 1])
    with btn_cols[0]:
        if st.button("🗑 删除所选", key="delete_btn", type="secondary", help="删除选中的国家"):
            df_now = data
            df_now = df_now[~df_now["del"]].reset_index(drop=True)
            st.session_state.edit_country_data = df_now
            st.rerun()
    with btn_cols[1]:
        st.write("")  # 空出空间

# 生成对比表格按钮
st.markdown("---")
submit_btn = st.button("📊 生成对比表格", type="primary")

if submit_btn and file_july and file_june and selected_commodity:
    with st.spinner("正在分析，请稍候..."):
        df_july = pd.read_csv(file_july)
        df_june = pd.read_csv(file_june)

        attribute = 'Production'
        year_new = 2025  # 25/26
        year_old = 2024  # 24/25

        results = []
        for row in data.itertuples():
            en_name = getattr(row, "en")
            cn_name = getattr(row, "cn")
            def get_pred(df, market_year, country):
                row_ = df[
                    (df['Commodity_Description'] == selected_commodity) &
                    (df['Attribute_Description'] == attribute) &
                    (df['Market_Year'] == market_year) &
                    (df['Country_Name'] == country)
                ]
                return float(row_['Value'].values[0]) if not row_.empty else ""
            jul_new = get_pred(df_july, year_new, en_name)
            jun_new = get_pred(df_june, year_new, en_name)
            jul_old = get_pred(df_july, year_old, en_name)
            mom = jul_new - jun_new if (jul_new != "" and jun_new != "") else ""
            yoy = jul_new - jul_old if (jul_new != "" and jul_old != "") else ""
            results.append({
                "国家": cn_name,
                "7月": jul_new,
                "6月": jun_new,
                "环比": mom,
                "24/25 7月": jul_old,
                "同比": yoy
            })
        df_out = pd.DataFrame(results)

        st.success(f"{cn_commodity} 产量对比已生成！下方可预览和下载Excel。")
        st.dataframe(df_out, use_container_width=True)

        # 高亮Excel
        def highlight_excel(df_out):
            output = io.BytesIO()
            df_out.to_excel(output, index=False)
            output.seek(0)
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            def highlight(ws, col_idx):
                for row in range(2, ws.max_row+1):
                    val = ws.cell(row, col_idx).value
                    if isinstance(val, (int, float)):
                        if val > 0:
                            ws.cell(row, col_idx).font = Font(color="008000")
                        elif val < 0:
                            ws.cell(row, col_idx).font = Font(color="FF0000")
            highlight(ws, 4) # 环比
            highlight(ws, 6) # 同比
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio

        st.download_button(
            f"⬇️ 下载{cn_commodity}产量对比Excel",
            data=highlight_excel(df_out),
            file_name=f"{cn_commodity}_output.xlsx"
        )

elif submit_btn:
    st.warning("请上传两个csv并选择作物种类")

st.markdown("""
---
**说明：**
- 你可以选择任何作物进行对比分析，也可手动输入或编辑作物/国家中文名。
- 国家英文名必须与csv内一致，否则无数据。
""")
