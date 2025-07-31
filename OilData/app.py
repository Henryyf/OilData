import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font

st.set_page_config(page_title="USDA大豆产量对比助手", layout="centered")

st.title("USDA大豆产量同比环比一键表格")

# 默认国家
default_countries = [
    {"en": "Brazil", "cn": "巴西"},
    {"en": "Argentina", "cn": "阿根廷"},
    {"en": "Paraguay", "cn": "巴拉圭"},
    {"en": "United States", "cn": "美国"},
    {"en": "China", "cn": "中国"}
]

with st.form("params"):
    st.markdown("#### 步骤一：上传最新两个月的csv（格式同USDA原表，无需修改）")
    col1, col2 = st.columns(2)
    file_july = col1.file_uploader("上传本月csv（如7月）", type="csv", key="july")
    file_june = col2.file_uploader("上传上月csv（如6月）", type="csv", key="june")

    st.markdown("#### 步骤二：如需修改国家名单（中英文都要填写）")
    data = st.data_editor(
        pd.DataFrame(default_countries), 
        use_container_width=True,
        column_order=["en", "cn"],
        key="edit_country"
    )
    submitted = st.form_submit_button("生成对比表格")

if submitted and file_july and file_june:
    with st.spinner("正在分析，请稍候..."):
        # 读取
        df_july = pd.read_csv(file_july)
        df_june = pd.read_csv(file_june)

        commodity = 'Oilseed, Soybean'
        attribute = 'Production'
        year_new = 2025  # 25/26
        year_old = 2024  # 24/25

        results = []
        for row in data.itertuples():
            en_name = getattr(row, "en")
            cn_name = getattr(row, "cn")
            def get_pred(df, market_year, country):
                row_ = df[
                    (df['Commodity_Description'] == commodity) &
                    (df['Attribute_Description'] == attribute) &
                    (df['Market_Year'] == market_year) &
                    (df['Country_Name'] == country)
                ]
                return float(row_['Value'].values[0]) if not row_.empty else ""
            jul_new = get_pred(df_july, year_new, en_name)
            jun_new = get_pred(df_june, year_new, en_name)
            jul_old = get_pred(df_july, year_old, en_name)
            mom = jul_new - jun_new if (jul_new != "" and jun_new != "") else ""
            yoy = jul_new - jul_old if (jul_new != "" and jul_old != "") else ""
            results.append({
                "国家": cn_name,
                "7月": jul_new,
                "6月": jun_new,
                "环比": mom,
                "24/25 7月": jul_old,
                "同比": yoy
            })
        df_out = pd.DataFrame(results)

        st.success("成功！下方可预览和下载Excel。")
        st.dataframe(df_out, use_container_width=True)

        # Excel高亮并导出
        def highlight_excel(df_out):
            output = io.BytesIO()
            df_out.to_excel(output, index=False)
            output.seek(0)
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            def highlight(ws, col_idx):
                for row in range(2, ws.max_row+1):
                    val = ws.cell(row, col_idx).value
                    if isinstance(val, (int, float)):
                        if val > 0:
                            ws.cell(row, col_idx).font = Font(color="008000")
                        elif val < 0:
                            ws.cell(row, col_idx).font = Font(color="FF0000")
            highlight(ws, 4) # 环比
            highlight(ws, 6) # 同比
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio

        st.download_button("下载Excel（含红绿高亮）", data=highlight_excel(df_out), file_name="usda_soybean_output.xlsx")

elif submitted:
    st.warning("请上传本月和上月的csv文件")

st.markdown("""
---
**温馨提示：**  
- 国家名必须是英文与原csv完全一致，否则抓取不到数值。
- 如果出现空白或报错，请检查csv数据、国家拼写和列名是否和原始USDA表格一致。
""")
