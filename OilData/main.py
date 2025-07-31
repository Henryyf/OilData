import pandas as pd
import openpyxl
from openpyxl.styles import Font

# 读取数据
df_july = pd.read_csv('psd_oilseeds_2025_July.csv')
df_june = pd.read_csv('psd_oilseeds_2025_June.csv')

commodity = 'Oilseed, Soybean'
attribute = 'Production'
year_new = 2025  # 25/26
year_old = 2024  # 24/25

# 固定国家列表
countries = [
    ("Brazil", "巴西"),
    ("Argentina", "阿根廷"),
    ("Paraguay", "巴拉圭"),
    ("United States", "美国"),
    ("China", "中国")
]

def get_pred(df, market_year, country):
    row = df[
        (df['Commodity_Description'] == commodity) &
        (df['Attribute_Description'] == attribute) &
        (df['Market_Year'] == market_year) &
        (df['Country_Name'] == country)
    ]
    return float(row['Value'].values[0]) if not row.empty else ""

results = []
for en_name, cn_name in countries:
    jul_new = get_pred(df_july, year_new, en_name)
    jun_new = get_pred(df_june, year_new, en_name)
    jul_old = get_pred(df_july, year_old, en_name)
    mom = jul_new - jun_new if (jul_new != "" and jun_new != "") else ""
    yoy = jul_new - jul_old if (jul_new != "" and jul_old != "") else ""
    results.append({
        "国家": cn_name,
        "7月": jul_new,
        "6月": jun_new,
        "环比": mom,
        "24/25 7月": jul_old,
        "同比": yoy
    })

df_out = pd.DataFrame(results)

# 输出Excel并高亮
excel_path = "usda_soybean_output.xlsx"
df_out.to_excel(excel_path, index=False)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active
def highlight(ws, col_idx):
    for row in range(2, ws.max_row+1):
        val = ws.cell(row, col_idx).value
        if isinstance(val, (int, float)):
            if val > 0:
                ws.cell(row, col_idx).font = Font(color="008000")  # 绿色
            elif val < 0:
                ws.cell(row, col_idx).font = Font(color="FF0000")  # 红色
highlight(ws, 4) # 环比
highlight(ws, 6) # 同比
wb.save(excel_path)

print("已生成Excel文件:", excel_path)
